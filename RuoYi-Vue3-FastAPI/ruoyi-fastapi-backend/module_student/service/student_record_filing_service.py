"""
学籍备案表 Service：导入、报告图生成（使用 report_fill_template2.py）
"""
import json
import os
import random
import re
import string
import subprocess
import tempfile
import zipfile
from datetime import date, datetime, timedelta
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
from sqlalchemy.ext.asyncio import AsyncSession

from config.env import UploadConfig
from exceptions.exception import ServiceException
from module_student.dao.student_record_filing_dao import StudentRecordFilingDao
from module_student.entity.do.student_record_filing_do import StudentRecordFiling
from utils.log_util import logger

COL_KEY_ALIASES = {
    "验证有效日期": ["验证有效日期", "验证有效期", "有效期"],
    "毕（结）业日期": ["毕（结）业日期", "毕(结)业日期", "毕业日期"],
    "毕（结）业": ["毕（结）业", "毕(结)业", "毕业", "结业"],
    "校（院）长姓名": ["校（院）长姓名", "校(院)长姓名", "校长姓名", "院长姓名"],
}
EXCEL_HEADERS_TEMPLATE2 = [
    "姓名", "性别", "出生日期", "入学日期", "毕（结）业日期", "学校名称", "专业", "学制",
    "层次", "学历类别", "学习形式", "毕（结）业", "证书编号", "校（院）长姓名", "验证有效日期", "照片",
]
EXCEL_HEADER_TO_DB_FILING = {
    "姓名": "name", "性别": "gender", "出生日期": "birth_date", "入学日期": "enrollment_date",
    "毕（结）业日期": "graduation_date", "学校名称": "school_name", "专业": "major",
    "学制": "duration", "层次": "level", "学历类别": "education_type", "学习形式": "learning_form",
    "毕（结）业": "graduation_status", "证书编号": "certificate_no", "校（院）长姓名": "president_name",
    "验证有效日期": "valid_until",
}


def gen_verification_code(length: int = 16) -> str:
    return "".join(random.choices(string.ascii_uppercase + string.digits, k=length))


class StudentRecordFilingService:
    @staticmethod
    def _backend_dir() -> str:
        return os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))

    @staticmethod
    def _report_template_path() -> str:
        return os.path.join(StudentRecordFilingService._backend_dir(), "uploads", "pic", "templete2.jpg")

    @staticmethod
    def _report_config_path() -> str:
        return os.path.join(StudentRecordFilingService._backend_dir(), "uploads", "pic", "layout_config2.json")

    @staticmethod
    def _photo_dir() -> str:
        return os.path.join(StudentRecordFilingService._backend_dir(), "uploads", "pic", "photo")

    @staticmethod
    def _report_fill_script_path() -> str:
        backend_dir = StudentRecordFilingService._backend_dir()
        project_root = os.path.abspath(os.path.join(backend_dir, "..", ".."))
        return os.path.join(project_root, "scripts", "report_fill_template2.py")

    @classmethod
    def _photo_file_path(cls, verification_code: str) -> str | None:
        if not verification_code or not verification_code.strip():
            return None
        code = verification_code.strip()
        d = cls._photo_dir()
        for ext in (".png", ".jpg", ".jpeg"):
            p = os.path.join(d, f"{code}{ext}")
            if os.path.isfile(p):
                return p
        return None

    @classmethod
    def save_photo_file(cls, verification_code: str, photo_bytes: bytes) -> str:
        code = (verification_code or "").strip()
        if not code:
            raise ValueError("验证码不能为空")
        photo_dir = cls._photo_dir()
        os.makedirs(photo_dir, exist_ok=True)
        path = os.path.abspath(os.path.join(photo_dir, f"{code}.png"))
        with open(path, "wb") as f:
            f.write(photo_bytes)
        return path

    @classmethod
    async def get_by_id(cls, db: AsyncSession, pk: int) -> StudentRecordFiling | None:
        return await StudentRecordFilingDao.get_by_id(db, pk)

    @classmethod
    async def get_by_code(cls, db: AsyncSession, code: str) -> StudentRecordFiling | None:
        return await StudentRecordFilingDao.get_by_verification_code(db, code)

    @classmethod
    async def list_students(cls, db: AsyncSession, query_object: dict | None = None, is_page: bool = True):
        return await StudentRecordFilingDao.get_list(db, query_object, is_page)

    @classmethod
    def _parse_valid_until(cls, val: Any) -> date | None:
        if val is None or (isinstance(val, str) and not val.strip()):
            return None
        try:
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, date):
                return val
            if isinstance(val, (int, float)) and 1 <= val <= 1000000:
                return date(1899, 12, 30) + timedelta(days=int(float(val)))
            s = str(val).strip()
            try:
                n = float(s)
                if 1000 <= n <= 100000:
                    return date(1899, 12, 30) + timedelta(days=int(n))
            except (ValueError, TypeError):
                pass
            m = re.match(r"(\d{4})年(\d{1,2})月(\d{1,2})日", s)
            if m:
                return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            return pd.to_datetime(val).date()
        except Exception:
            return None

    @classmethod
    def _format_date_for_storage(cls, val: Any) -> str:
        parsed = cls._parse_valid_until(val)
        return parsed.strftime("%Y年%m月%d日") if parsed else (str(val).strip() if val else "")

    @classmethod
    def _build_col_index(cls, header_row: tuple) -> dict:
        col_index = {}
        col_norm = {}
        for idx, v in enumerate(header_row):
            if v is None:
                continue
            k = str(v).strip()
            if not k:
                continue
            col_index[k] = idx
            norm = k.replace(" ", "").replace("　", "")
            col_norm[norm] = k
        for norm, orig in col_norm.items():
            if norm not in col_index:
                col_index[norm] = col_index[orig]
        return col_index

    @classmethod
    def _find_col(cls, col_index: dict, *keys: str) -> int | None:
        for k in keys:
            if k in col_index:
                return col_index[k]
            norm = k.replace(" ", "").replace("　", "")
            if norm in col_index:
                return col_index[norm]
        return None

    @classmethod
    async def _save_one_row(
        cls,
        db: AsyncSession,
        get_cell: Any,
        valid_until: date,
        name: str,
        photo_blob: bytes | None,
        create_by: str,
    ) -> None:
        code = gen_verification_code()
        while await StudentRecordFilingDao.exists_verification_code(db, code):
            code = gen_verification_code()
        def _s(k: str) -> str:
            v = get_cell(k)
            return "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v).strip()
        date_fmt = cls._format_date_for_storage
        data = {
            "verification_code": code,
            "update_date": date_fmt(get_cell("更新日期")) or datetime.now().strftime("%Y年%m月%d日"),
            "name": name,
            "gender": _s("性别"),
            "birth_date": date_fmt(get_cell("出生日期")) or _s("出生日期"),
            "enrollment_date": date_fmt(get_cell("入学日期")) or _s("入学日期"),
            "graduation_date": date_fmt(get_cell("毕（结）业日期")) or _s("毕（结）业日期"),
            "school_name": _s("学校名称"),
            "major": _s("专业"),
            "duration": _s("学制"),
            "level": _s("层次"),
            "education_type": _s("学历类别"),
            "learning_form": _s("学习形式"),
            "graduation_status": _s("毕（结）业"),
            "certificate_no": _s("证书编号"),
            "president_name": _s("校（院）长姓名"),
            "valid_until": valid_until,
            "photo_blob": photo_blob,
            "del_flag": "0",
            "create_by": create_by or "",
        }
        await StudentRecordFilingDao.add(db, data)
        if photo_blob and code:
            try:
                os.makedirs(cls._photo_dir(), exist_ok=True)
                p = os.path.join(cls._photo_dir(), f"{code}.png")
                with open(p, "wb") as f:
                    f.write(photo_blob)
            except Exception as e:
                logger.warning("保存照片文件失败 %s: %s", code, e)

    @classmethod
    async def _import_from_xlsx(cls, db: AsyncSession, file_path: str, create_by: str) -> dict:
        wb = load_workbook(file_path, read_only=False, data_only=False)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header_row = tuple(header_row) if header_row else ()
        if not header_row:
            wb.close()
            raise ServiceException(message="Excel 无表头")
        col_index = cls._build_col_index(header_row)
        name_col = cls._find_col(col_index, "姓名")
        date_col = cls._find_col(col_index, "验证有效日期", "验证有效期", "有效期")
        if name_col is None or date_col is None:
            wb.close()
            raise ServiceException(message="表头需含「姓名」「验证有效日期」列")
        row_to_photo = {}
        if hasattr(ws, "_images"):
            for img in ws._images:
                anchor = getattr(img, "anchor", None)
                if anchor is None:
                    continue
                row_1based = None
                if isinstance(anchor, str):
                    try:
                        col_letter, row = coordinate_from_string(anchor)
                        row_1based = int(row)
                    except Exception:
                        continue
                else:
                    from_obj = getattr(anchor, "_from", None)
                    if from_obj is not None:
                        row_0based = getattr(from_obj, "row", None)
                        if row_0based is not None:
                            row_1based = int(row_0based) + 1
                if row_1based is not None:
                    try:
                        row_to_photo[row_1based] = img._data()
                    except Exception:
                        pass
        if not row_to_photo:
            try:
                with zipfile.ZipFile(file_path, "r") as zf:
                    media_files = sorted([n for n in zf.namelist() if n.startswith("xl/media/")])
                    for i, name in enumerate(media_files):
                        if name.lower().endswith((".png", ".jpg", ".jpeg", ".gif")):
                            row_to_photo[2 + i] = zf.read(name)
            except Exception:
                pass

        def _cell(row: tuple, key: str, default=None):
            keys = COL_KEY_ALIASES.get(key, [key])
            idx = cls._find_col(col_index, *keys)
            if idx is None:
                return default
            if idx >= len(row):
                return default
            return row[idx]

        success, fail = 0, 0
        fail_details = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row = tuple(row) if row else ()
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            valid_until = cls._parse_valid_until(_cell(row, "验证有效日期"))
            if valid_until is None:
                fail += 1
                fail_details.append(f"第{row_idx}行：验证有效日期为空或格式错误")
                continue
            name = _cell(row, "姓名") or ""
            if not str(name).strip():
                fail += 1
                fail_details.append(f"第{row_idx}行：姓名为空")
                continue
            photo_blob = row_to_photo.get(row_idx)
            get_cell = lambda k, r=row: _cell(r, k)
            await cls._save_one_row(db, get_cell, valid_until, str(name).strip(), photo_blob, create_by)
            success += 1
        wb.close()
        return {"success": success, "fail": fail, "fail_details": fail_details}

    @classmethod
    async def _import_from_xls(cls, db: AsyncSession, file_path: str, create_by: str) -> dict:
        try:
            df = pd.read_excel(file_path, engine="xlrd", header=0)
        except Exception as e:
            raise ServiceException(message=f"读取 .xls 失败：{e}") from e
        if df.empty:
            raise ServiceException(message="Excel 无有效数据")
        col_map = {str(c).strip(): c for c in df.columns if str(c).strip()}
        col_map_norm = {k.replace(" ", "").replace("　", ""): k for k in col_map}
        def _col(key):
            if key in col_map:
                return col_map[key]
            k = key.replace(" ", "").replace("　", "")
            return col_map.get(col_map_norm.get(k))
        if not _col("姓名") or not _col("验证有效日期"):
            raise ServiceException(message="表头需含「姓名」「验证有效日期」")

        def _val(row, key, default=None):
            col = _col(key)
            if not col:
                return default
            v = row.get(col)
            return default if pd.isna(v) else v

        success, fail = 0, 0
        fail_details = []
        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            if row.isna().all():
                continue
            valid_until = cls._parse_valid_until(_val(row, "验证有效日期"))
            if valid_until is None:
                fail += 1
                fail_details.append(f"第{row_idx}行：验证有效日期格式错误")
                continue
            name = _val(row, "姓名") or ""
            if not str(name).strip():
                fail += 1
                fail_details.append(f"第{row_idx}行：姓名为空")
                continue
            get_cell = lambda k: _val(row, k)
            await cls._save_one_row(db, get_cell, valid_until, str(name).strip(), None, create_by)
            success += 1
        return {"success": success, "fail": fail, "fail_details": fail_details}

    @classmethod
    async def import_from_excel(cls, db: AsyncSession, file_path: str, create_by: str = "") -> dict:
        if not os.path.isfile(file_path):
            raise ServiceException(message="文件不存在")
        ext = file_path.lower()
        if ext.endswith(".xls") and not ext.endswith(".xlsx"):
            return await cls._import_from_xls(db, file_path, create_by)
        return await cls._import_from_xlsx(db, file_path, create_by)

    @classmethod
    async def update_student(cls, db: AsyncSession, pk: int, data: dict, update_by: str = "") -> None:
        obj = await StudentRecordFilingDao.get_by_id(db, pk)
        if not obj:
            raise ServiceException(message="记录不存在")
        update_data = dict(data)
        if "valid_until" in update_data and update_data["valid_until"] is not None:
            parsed = cls._parse_valid_until(update_data["valid_until"])
            if parsed is None:
                raise ServiceException(message="验证有效日期格式不正确")
            update_data["valid_until"] = parsed
        for f in ("birth_date", "enrollment_date", "graduation_date", "update_date"):
            if f in update_data and update_data[f] is not None and update_data[f] != "":
                update_data[f] = cls._format_date_for_storage(update_data[f])
        if update_by:
            update_data["update_by"] = update_by
        update_data["update_time"] = datetime.now()
        await StudentRecordFilingDao.update_by_id(db, pk, update_data)

    @classmethod
    async def update_photo_blob(cls, db: AsyncSession, student_id: int, photo_bytes: bytes) -> None:
        await StudentRecordFilingDao.update_by_id(db, student_id, {"photo_blob": photo_bytes})

    @classmethod
    def _format_date_cn(cls, val: Any) -> str:
        if val is None or (isinstance(val, str) and not val.strip()):
            return ""
        try:
            if isinstance(val, (datetime, date)):
                return val.strftime("%Y年%m月%d日")
            if isinstance(val, (int, float)) and 1 <= val <= 1000000:
                d = date(1899, 12, 30) + timedelta(days=int(float(val)))
                return d.strftime("%Y年%m月%d日")
            s = str(val).strip()
            if re.match(r"^\d{4}年\d{1,2}月\d{1,2}日", s):
                return s
            parsed = cls._parse_valid_until(val)
            if parsed:
                return parsed.strftime("%Y年%m月%d日")
        except Exception:
            pass
        return str(val)

    @classmethod
    def render_report_image(
        cls,
        student: StudentRecordFiling,
        verify_base_url: str,
        output_path: str | None = None,
    ) -> bytes:
        """使用 report_fill_template2.py 生成备案表报告图"""
        template_path = cls._report_template_path()
        config_path = cls._report_config_path()
        script_path = cls._report_fill_script_path()
        if not os.path.isfile(template_path):
            raise ServiceException(message="备案表模板不存在")
        if not os.path.isfile(config_path):
            raise ServiceException(message="布局配置不存在")
        if not os.path.isfile(script_path):
            raise ServiceException(message="报告填充脚本不存在")
        fmt = cls._format_date_cn
        data = {
            "更新日期": fmt(student.update_date) or datetime.now().strftime("%Y年%m月%d日"),
            "姓名": student.name or "",
            "性别": student.gender or "",
            "出生日期": fmt(student.birth_date),
            "入学日期": fmt(student.enrollment_date),
            "毕（结）业日期": fmt(student.graduation_date),
            "学校名称": student.school_name or "",
            "专业": student.major or "",
            "学制": student.duration or "",
            "层次": student.level or "",
            "学历类别": student.education_type or "",
            "学习形式": student.learning_form or "",
            "毕（结）业": student.graduation_status or "",
            "证书编号": student.certificate_no or "",
            "校（院）长姓名": student.president_name or "",
            "在线验证码": student.verification_code or "",
            "二维码内容": f"{verify_base_url.rstrip('/')}/verify?code={student.verification_code}",
        }
        with tempfile.TemporaryDirectory() as tmp:
            data_json = os.path.join(tmp, "data.json")
            with open(data_json, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False)
            photo_path = None
            if student.photo_blob:
                photo_path = os.path.join(tmp, "photo.png")
                with open(photo_path, "wb") as f:
                    f.write(student.photo_blob)
            else:
                fp = cls._photo_file_path(student.verification_code or "")
                if fp and os.path.isfile(fp):
                    photo_path = fp
            out_path = output_path or os.path.join(tmp, "result.png")
            cmd = [
                "python",
                script_path,
                "--template", template_path,
                "--config", config_path,
                "--output", out_path,
                "--data-json", data_json,
            ]
            if photo_path and os.path.isfile(photo_path):
                cmd.extend(["--photo", photo_path])
            try:
                subprocess.run(cmd, check=True, capture_output=True, timeout=60, cwd=os.path.dirname(script_path))
            except subprocess.CalledProcessError as e:
                logger.exception("report_fill_template2 failed: %s", e.stderr)
                raise ServiceException(message="生成备案表报告图失败")
            except Exception as e:
                logger.exception("render_report_image: %s", e)
                raise ServiceException(message="生成备案表报告图失败")
            with open(out_path, "rb") as f:
                return f.read()
